"""Helpers puros para preparar datos reales al formato de la planta modelo.

Convierten/alinean dos series horarias —CMg del Coordinador y generación PV del
Explorador Solar— en el CSV `timestamp,generacion_w,cmg_mills_por_mwh` que consume
`GatewayCSV`. Todo es función pura y testeable; el HTTP vive en la CLI.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

Serie = list[tuple[str, int]]

_EXTENSIONES_XLSX = {".xlsx", ".xlsm"}


def parsear_decimal(texto: str) -> float:
    """Convierte un número en texto a float, tolerando la coma decimal chilena.

    - ``"57,79415"`` -> 57.79415  (coma decimal)
    - ``"1.234,56"`` -> 1234.56   (punto de miles + coma decimal)
    - ``"78500.4"`` -> 78500.4    (punto decimal estándar)
    - ``"80000"`` -> 80000.0

    Regla: si hay coma, la coma es el separador decimal y los puntos son de miles;
    si no hay coma, el punto es decimal. (Exporta sin separador de miles para evitar
    ambigüedad en valores como ``"80.000"``.)
    """
    valor = (texto or "").strip()
    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")
    return float(valor)


def leer_serie_csv(
    ruta: Path | str,
    columna_ts: str,
    columna_valor: str,
    escala: float = 1.0,
    fila_encabezado: int = 1,
) -> Serie:
    """Lee un CSV como serie ``(timestamp, valor_entero)``, escalando el valor.

    ``escala`` convierte unidades (p. ej. 1000 si la generación viene en kW y se
    quiere en W). El valor se redondea a entero. ``fila_encabezado`` (1-based) permite
    saltar filas de metadatos sobre la tabla (el export del Explorador Solar trae ~54).
    Filas con timestamp o valor vacíos al final se omiten.
    """
    ruta = Path(ruta)
    with ruta.open(newline="", encoding="utf-8") as archivo:
        for _ in range(fila_encabezado - 1):
            next(archivo, None)
        lector = csv.DictReader(archivo)
        faltantes = {columna_ts, columna_valor} - set(lector.fieldnames or [])
        if faltantes:
            raise ValueError(f"Faltan columnas en {ruta.name}: {sorted(faltantes)}")
        serie: Serie = []
        for numero_fila, fila in enumerate(lector, fila_encabezado + 1):
            timestamp = (fila[columna_ts] or "").strip()
            crudo = (fila[columna_valor] or "").strip()
            if not timestamp and not crudo:
                continue  # fila en blanco (típica al final del archivo)
            if not timestamp:
                raise ValueError(f"Fila {numero_fila}: timestamp vacío")
            try:
                valor = round(parsear_decimal(crudo) * escala)
            except ValueError as error:
                raise ValueError(f"Fila {numero_fila}: valor inválido ({error})") from error
            serie.append((timestamp, valor))
    return serie


def _celda_a_texto_ts(valor: Any) -> str:
    """Normaliza una celda de timestamp a texto ISO (datetime nativo o string)."""
    if isinstance(valor, datetime):
        return valor.isoformat(timespec="minutes")
    if isinstance(valor, date):
        return valor.isoformat()
    return str(valor or "").strip()


def _fecha_texto(valor: Any) -> str:
    """Extrae la parte de fecha (YYYY-MM-DD) de una celda datetime, date o texto."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    texto = str(valor or "").strip()
    return texto.split("T")[0].split(" ")[0]


def _celda_a_decimal(valor: Any) -> float:
    """Convierte una celda numérica o de texto a float (tolera coma chilena)."""
    if isinstance(valor, bool):  # bool es subclase de int; no es un número válido aquí
        raise ValueError(f"valor booleano inesperado: {valor!r}")
    if isinstance(valor, int | float):
        return float(valor)
    return parsear_decimal(str(valor or ""))


def _normalizar_columna(nombre: str) -> str:
    return nombre.strip().casefold().replace(" ", "").replace("_", "").replace(".", "")


def _indice_columna(encabezado: list[str], nombre: str, archivo: str) -> int:
    """Ubica una columna con matching tolerante (para nombres como ``S.GREGORIO____013``).

    Prueba, en orden: coincidencia exacta, normalizada (sin espacios/guiones/puntos,
    case-insensitive) y por prefijo normalizado. Exige que sea única.
    """
    if nombre in encabezado:
        return encabezado.index(nombre)
    objetivo = _normalizar_columna(nombre)
    normal = [_normalizar_columna(c) for c in encabezado]
    exactos = [i for i, c in enumerate(normal) if c == objetivo]
    if len(exactos) == 1:
        return exactos[0]
    if not exactos and objetivo:
        prefijos = [i for i, c in enumerate(normal) if c.startswith(objetivo)]
        if len(prefijos) == 1:
            return prefijos[0]
    raise ValueError(
        f"Columna '{nombre}' no encontrada de forma única en {archivo}. "
        f"Columnas: {[c for c in encabezado if c]}"
    )


def leer_serie_xlsx(
    ruta: Path | str,
    columna_ts: str,
    columna_valor: str,
    escala: float = 1.0,
    hoja: str | None = None,
    fila_encabezado: int = 1,
    columna_hora: str | None = None,
) -> Serie:
    """Lee una hoja .xlsx como serie ``(timestamp, valor_entero)``.

    Mismo contrato que `leer_serie_csv` pero sobre Excel: las descargas del
    Coordinador y del Explorador Solar vienen en .xlsx. ``fila_encabezado`` (1-based)
    permite saltar filas de metadatos sobre la tabla; ``hoja`` elige la pestaña.
    Las celdas pueden venir como número/fecha nativos o como texto con coma decimal.
    El nombre de columna se resuelve con matching tolerante (ver `_indice_columna`).

    **Formato ancho del Coordinador:** cuando el timestamp está partido en ``Fecha``
    (celda combinada por día) + ``Hora`` (0..23), pasa ``columna_hora``: la fecha se
    arrastra hacia abajo (forward-fill de la celda combinada) y se combina con la hora.
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:  # pragma: no cover - depende del entorno
        raise RuntimeError(
            "Leer .xlsx requiere openpyxl: instala el extra `acopia[ingesta]`"
        ) from error

    ruta = Path(ruta)
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        hoja_obj = libro[hoja] if hoja is not None else libro.active
        filas = hoja_obj.iter_rows(min_row=fila_encabezado, values_only=True)
        try:
            encabezado = [str(c).strip() if c is not None else "" for c in next(filas)]
        except StopIteration as error:
            raise ValueError(f"{ruta.name}: hoja vacía") from error
        i_ts = _indice_columna(encabezado, columna_ts, ruta.name)
        i_valor = _indice_columna(encabezado, columna_valor, ruta.name)
        i_hora = _indice_columna(encabezado, columna_hora, ruta.name) if columna_hora else None

        serie: Serie = []
        ultima_fecha = ""  # forward-fill de la fecha (celda combinada por día)
        for numero_fila, fila in enumerate(filas, fila_encabezado + 1):
            if i_hora is None:
                timestamp = _celda_a_texto_ts(fila[i_ts])
                if not timestamp:
                    continue  # filas en blanco al final de la hoja
            else:
                if fila[i_ts] is not None and str(fila[i_ts]).strip():
                    ultima_fecha = _fecha_texto(fila[i_ts])
                celda_hora = fila[i_hora]
                if celda_hora is None or str(celda_hora).strip() == "" or not ultima_fecha:
                    continue  # fila en blanco o antes de la primera fecha
                try:
                    hora = int(_celda_a_decimal(celda_hora))
                except ValueError as error:
                    raise ValueError(f"Fila {numero_fila}: hora inválida ({error})") from error
                timestamp = f"{ultima_fecha}T{hora:02d}:00"
            try:
                valor = round(_celda_a_decimal(fila[i_valor]) * escala)
            except ValueError as error:
                raise ValueError(f"Fila {numero_fila}: valor inválido ({error})") from error
            serie.append((timestamp, valor))
    finally:
        libro.close()
    return serie


def leer_serie(
    ruta: Path | str,
    columna_ts: str,
    columna_valor: str,
    escala: float = 1.0,
    hoja: str | None = None,
    fila_encabezado: int = 1,
    columna_hora: str | None = None,
) -> Serie:
    """Lee una serie despachando por extensión: .xlsx/.xlsm -> Excel, resto -> CSV."""
    if Path(ruta).suffix.lower() in _EXTENSIONES_XLSX:
        return leer_serie_xlsx(
            ruta, columna_ts, columna_valor, escala, hoja, fila_encabezado, columna_hora
        )
    if columna_hora:
        raise ValueError("columna_hora solo se soporta en .xlsx (formato ancho del Coordinador)")
    return leer_serie_csv(ruta, columna_ts, columna_valor, escala, fila_encabezado)


def extraer_cmg(
    resultados: Iterable[dict[str, Any]],
    campo_ts: str,
    campo_valor: str,
    escala: float = 1.0,
) -> Serie:
    """Extrae ``(timestamp, cmg)`` de los registros JSON de la API del Coordinador."""
    serie: Serie = []
    for registro in resultados:
        timestamp = str(registro[campo_ts])
        valor = round(parsear_decimal(str(registro[campo_valor])) * escala)
        serie.append((timestamp, valor))
    return serie


def alinear_series(generacion: Serie, cmg: Serie) -> list[tuple[str, int, int]]:
    """Cruza generación y CMg por timestamp (inner join), ordenado cronológicamente.

    Solo conserva los timestamps presentes en ambas series.
    """
    cmg_por_ts = dict(cmg)
    filas = [
        (timestamp, valor_gen, cmg_por_ts[timestamp])
        for timestamp, valor_gen in generacion
        if timestamp in cmg_por_ts
    ]
    return sorted(filas)


def alinear_por_posicion(generacion: Serie, cmg: Serie) -> list[tuple[str, int, int]]:
    """Cruza generación y CMg por **posición** (hora a hora), usando el timestamp del CMg.

    Útil cuando las series son de años distintos (p. ej. CMg real reciente + generación
    de un "año típico" del Explorador Solar): no comparten calendario, pero ambas son
    horarias y se aparean por índice. Exige el mismo largo para evitar desfases silenciosos.
    """
    if len(generacion) != len(cmg):
        raise ValueError(
            f"Las series deben tener el mismo largo para alinear por posición: "
            f"generación={len(generacion)}, cmg={len(cmg)}"
        )
    return [
        (cmg_ts, valor_gen, valor_cmg)
        for (_, valor_gen), (cmg_ts, valor_cmg) in zip(generacion, cmg, strict=True)
    ]


def escribir_csv_planta(ruta: Path | str, filas: Sequence[tuple[str, int, int]]) -> None:
    """Escribe el CSV de planta que consume `GatewayCSV`."""
    with Path(ruta).open("w", newline="", encoding="utf-8") as archivo:
        escritor = csv.writer(archivo)
        escritor.writerow(["timestamp", "generacion_w", "cmg_mills_por_mwh"])
        escritor.writerows(filas)


def escribir_serie_csv(ruta: Path | str, serie: Serie, nombre_valor: str) -> None:
    """Escribe una serie ``(timestamp, valor)`` como CSV de dos columnas."""
    with Path(ruta).open("w", newline="", encoding="utf-8") as archivo:
        escritor = csv.writer(archivo)
        escritor.writerow(["timestamp", nombre_valor])
        escritor.writerows(serie)
