"""Lector del XLSX "Reducciones ERV" del Coordinador (vertimiento renovable).

El Coordinador publica mensualmente "Reducciones de Energía Eólica-Solar-Hidro en
el SEN" (ADR-012): un libro con una hoja ``Resumen-DiarioHorario-<Tecnología>`` por
tecnología, donde cada día es un bloque — una fila con la fecha, una fila de
encabezado ``Central/Hora`` con las horas 1..24, una fila por central con la energía
reducida en **MWh**, y una fila ``Total`` que cierra el bloque. Este módulo lo
aplana a registros ``(tecnologia, central, fecha, 24 valores)``; los agregados
viven en `acopia.interfaces.observatorio`.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from acopia.infrastructure.ingesta.preparacion import _celda_a_decimal, _fecha_texto

_PREFIJO_HOJA = "Resumen-DiarioHorario-"
_TECNOLOGIAS_POR_SUFIJO = {
    "eolico": "eolica",
    "solar": "solar",
    "hp": "hidro_pasada",
    "he": "hidro_embalse",
}
_HORAS = 24
_MINIMO_HORAS_REPARABLE = 20  # repara typos puntuales del encabezado, no headers truncos


@dataclass(frozen=True)
class ReduccionDiaria:
    """Vertimiento de una central en un día: 24 valores horarios en MWh."""

    tecnologia: str
    central: str
    fecha: str  # YYYY-MM-DD
    energia_mwh: tuple[float, ...]  # índice 0 = hora 1 del Coordinador

    @property
    def total_mwh(self) -> float:
        return sum(self.energia_mwh)


def _sin_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c)
    )


def _tecnologia_de_hoja(nombre_hoja: str) -> str | None:
    """Mapea ``Resumen-DiarioHorario-Eólico`` -> ``eolica`` (None si no es una hoja ERV)."""
    if not nombre_hoja.startswith(_PREFIJO_HOJA):
        return None
    sufijo = _sin_acentos(nombre_hoja.removeprefix(_PREFIJO_HOJA)).casefold()
    return _TECNOLOGIAS_POR_SUFIJO.get(sufijo)


def _es_fecha(celda: Any) -> bool:
    texto = _fecha_texto(celda)
    return len(texto) == 10 and texto[4] == "-" and texto[7] == "-"


def _indices_horas(fila: tuple[Any, ...], numero_fila: int, hoja: str) -> tuple[int, ...]:
    """Ubica las 24 columnas horarias en la fila ``Central/Hora`` (encabezados 1..24).

    El archivo real trae typos en el encabezado (mayo 2026: un ``|`` donde iba el
    ``2``), pero las horas ocupan columnas contiguas; si toda hora encontrada calza
    con ``col(h) = col(1) + (h-1)``, las faltantes se reparan por posición. Si el
    patrón no calza, se falla: mejor un error que un desfase silencioso de horas.
    """
    indices: dict[int, int] = {}
    for i, celda in enumerate(fila):
        if isinstance(celda, int | float) and not isinstance(celda, bool):
            hora = int(celda)
            if 1 <= hora <= _HORAS and hora == celda and hora not in indices:
                indices[hora] = i
    if len(indices) < _MINIMO_HORAS_REPARABLE:
        raise ValueError(
            f"{hoja}, fila {numero_fila}: el encabezado 'Central/Hora' no trae las "
            f"24 horas (encontradas: {len(indices)}; con menos de "
            f"{_MINIMO_HORAS_REPARABLE} no se repara por posición)"
        )
    hora_ancla, i_ancla = next(iter(indices.items()))
    i_hora_1 = i_ancla - (hora_ancla - 1)
    if any(i != i_hora_1 + (hora - 1) for hora, i in indices.items()):
        raise ValueError(
            f"{hoja}, fila {numero_fila}: las horas del encabezado no ocupan columnas "
            f"contiguas; no se puede reparar el typo por posición"
        )
    return tuple(i_hora_1 + (hora - 1) for hora in range(1, _HORAS + 1))


def leer_reducciones_erv(ruta: Path | str) -> tuple[ReduccionDiaria, ...]:
    """Aplana las hojas diario/horario del XLSX de Reducciones ERV.

    Devuelve un registro por (tecnología, central, día) con los 24 MWh horarios.
    Celdas vacías cuentan como 0.0 (la matriz real las trae en cero explícito, pero
    el formato no está garantizado). Valores negativos son error: el vertimiento es
    energía no inyectada, no puede ser negativo. Filas ``Total`` se omiten (los
    totales se recalculan aguas arriba — nunca se confía en fórmulas del libro).
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:  # pragma: no cover - depende del entorno
        raise RuntimeError(
            "Leer .xlsx requiere openpyxl: instala el extra `acopia[ingesta]`"
        ) from error

    ruta = Path(ruta)
    libro = load_workbook(ruta, read_only=True, data_only=True)
    try:
        registros: list[ReduccionDiaria] = []
        hojas_erv = 0
        for hoja_obj in libro.worksheets:
            tecnologia = _tecnologia_de_hoja(hoja_obj.title)
            if tecnologia is None:
                continue
            hojas_erv += 1
            registros.extend(_leer_hoja(hoja_obj, tecnologia))
        if hojas_erv == 0:
            raise ValueError(
                f"{ruta.name}: ninguna hoja '{_PREFIJO_HOJA}*'; ¿es el XLSX de "
                f"Reducciones ERV del Coordinador?"
            )
    finally:
        libro.close()
    return tuple(registros)


def _leer_hoja(hoja_obj: Any, tecnologia: str) -> list[ReduccionDiaria]:
    hoja = str(hoja_obj.title)
    registros: list[ReduccionDiaria] = []
    fecha = ""
    indices: tuple[int, ...] | None = None
    for numero_fila, fila in enumerate(hoja_obj.iter_rows(values_only=True), 1):
        celdas = [c for c in fila if c is not None and str(c).strip() != ""]
        if not celdas:
            continue
        primera = celdas[0]
        if _es_fecha(primera):
            fecha = _fecha_texto(primera)  # abre el bloque del día
            indices = None
            continue
        etiqueta = str(primera).strip()
        if etiqueta.casefold().replace(" ", "") == "central/hora":
            indices = _indices_horas(fila, numero_fila, hoja)
            continue
        if not fecha or indices is None or etiqueta.casefold() == "total":
            continue  # títulos del libro, o la fila Total que cierra el bloque
        energia = _energia_de_fila(fila, indices, numero_fila, hoja)
        registros.append(ReduccionDiaria(tecnologia, etiqueta, fecha, energia))
    return registros


def _energia_de_fila(
    fila: tuple[Any, ...], indices: tuple[int, ...], numero_fila: int, hoja: str
) -> tuple[float, ...]:
    energia: list[float] = []
    for hora, i in enumerate(indices, 1):
        celda = fila[i] if i < len(fila) else None
        # el archivo real trae "-" (guión contable) por "sin reducción" en algunas horas
        if celda is None or str(celda).strip() in {"", "-"}:
            energia.append(0.0)
            continue
        try:
            valor = _celda_a_decimal(celda)
        except ValueError as error:
            raise ValueError(
                f"{hoja}, fila {numero_fila}, hora {hora}: valor inválido ({error})"
            ) from error
        if valor < 0:
            raise ValueError(
                f"{hoja}, fila {numero_fila}, hora {hora}: vertimiento negativo ({valor})"
            )
        energia.append(valor)
    return tuple(energia)
